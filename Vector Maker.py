from openpyxl import Workbook, load_workbook

spywb = load_workbook(r'C:\Users\KRAZY\Documents\Clean Data\SPYDATA.xlsx', data_only=True)
spy = spywb['SDAT']

wb = Workbook()
ws =wb.active

ini = 32
for w in range(7628):
    date = spy["A"+str(ini)].value
    i = int(spy["E"+str(ini)].value)
    ii = int(spy["E"+str(ini-1)].value)
    iii = int(spy["E"+str(ini-2)].value)
    iv = int(spy["E"+str(ini-3)].value)
    v = int(spy["E"+str(ini-4)].value)
    vi = int(spy["E"+str(ini-5)].value)
    vii = int(spy["E"+str(ini-6)].value)
    viii = int(spy["E"+str(ini-7)].value)
    ix = int(spy["E"+str(ini-8)].value)
    x = int(spy["E"+str(ini-9)].value)
    xi = int(spy["E"+str(ini-10)].value)
    xii = int(spy["E"+str(ini-11)].value)
    xiii = int(spy["E"+str(ini-12)].value)
    xiv = int(spy["E"+str(ini-13)].value)
    xv = int(spy["E"+str(ini-14)].value)
    xvi = int(spy["E"+str(ini-15)].value)
    xvii = int(spy["E"+str(ini-16)].value)
    xviii = int(spy["E"+str(ini-17)].value)
    xix = int(spy["E"+str(ini-18)].value)
    xx = int(spy["E"+str(ini-19)].value)
    xxi = int(spy["E"+str(ini-20)].value)
    xxii = int(spy["E"+str(ini-21)].value)
    xxiii = int(spy["E"+str(ini-22)].value)
    xxiv = int(spy["E"+str(ini-23)].value)
    xxv = int(spy["E"+str(ini-24)].value)
    xxvi = int(spy["E"+str(ini-25)].value)
    xxvii = int(spy["E"+str(ini-26)].value)
    xxviii = int(spy["E"+str(ini-27)].value)
    xxix = int(spy["E"+str(ini-28)].value)
    xxx = int(spy["E"+str(ini-29)].value)
    xxxi = int(spy["E" + str(ini - 30)].value)

    ai = (i/ii)-1
    aii = (ii/iii) - 1
    aiii = (iii / iv) - 1
    aiv = (iv / v) - 1
    av = (v / vi) - 1
    avi = (vi / vii) - 1
    avii = (vii / viii) - 1
    aviii = (viii / ix) - 1
    aix = (ix / x) - 1
    ax = (x / xi) - 1
    axi = (xi / xii) - 1
    axii = (xii / xiii) - 1
    axiii = (xii / xiv) - 1
    axiv = (xiv / xv) - 1
    axv = (xv / xvi) - 1
    axvi = (xvi / xvii) - 1
    axvii = (xvii / xviii) - 1
    axviii = (xviii / xix) - 1
    axix = (xix / xx) - 1
    axx = (xx / xxi) - 1
    axxi = (xxi / xxii) - 1
    axxii = (xxii / xxiii) - 1
    axxiii = (xxiii / xxiv) - 1
    axxiv = (xxiv / xxv) - 1
    axxv = (xxv / xxvi) - 1
    axxvi = (xxvi / xxvii) - 1
    axxvii = (xxvii / xxviii) - 1
    axxviii = (xxviii / xxiv) - 1
    axxiv = (xxiv / xxv) - 1
    axxv = (xxv / xxvi) - 1
    axxvi = (xxvi / xxvii) - 1
    axxvii = (xxvii / xxviii) - 1
    axxviii = (xxviii / xxix) - 1
    axxix = (xxix / xxx) - 1
    axxx = (xxx / xxxi) - 1


    ws.append([date,ai,aii,aiii,aiv,av,avi,avii,aviii,aix,ax,axi,axii,axiii,axiv,axv,axvi,axvii,axviii,axix,axx,axxi,axxii,axxiii,axxiv,axxv,axxvi,axxvii,axxviii,axxix,axxx])

    ini = ini + 1
    print(ini)

wb.save(r'SPYVECTORSs.xlsx')